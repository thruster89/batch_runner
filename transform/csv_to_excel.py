import re
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from pandas.api.types import is_integer_dtype, is_float_dtype
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from util.paths import CSV_DIR, EXCEL_DIR, SQL_DIR


# ---------------------------------------------------------
# Excel 파일명 생성
# ---------------------------------------------------------
def get_excel_output_path(
    schema: str,
    top_folder: str,
    max_files: int = 10,
) -> Path:
    EXCEL_DIR.mkdir(exist_ok=True)

    ts = datetime.now().strftime("%y%m%d")
    base = f"{schema}_{top_folder}_{ts}"

    pattern = re.compile(rf"{base}_(\d+)\.xlsx")
    existing = []

    for f in EXCEL_DIR.glob(f"{base}_*.xlsx"):
        m = pattern.fullmatch(f.name)
        if m:
            existing.append((int(m.group(1)), f))

    next_idx = max((i for i, _ in existing), default=0) + 1
    out = EXCEL_DIR / f"{base}_{next_idx}.xlsx"

    existing_sorted = sorted(existing, key=lambda x: x[1].stat().st_mtime)
    while len(existing_sorted) >= max_files:
        _, old = existing_sorted.pop(0)
        old.unlink()
        logging.info("Old Excel file removed: %s", old.name)

    return out


# ---------------------------------------------------------
# CSV → Excel
# ---------------------------------------------------------
def csv_to_excel(source: str, host_name: str, schema: str, sql_files: list[Path]):
    """
    이번 실행 대상 SQL 기준으로 생성된 CSV만 Excel로 변환
    """

    base_dir = CSV_DIR / source / host_name

    csv_files = csv_files_from_sql(source, host_name, sql_files)

    if not csv_files:
        logging.info(
            "Excel export skipped | schema=%s | no CSV files",
            schema,
        )
        return

    grouped: dict[str, list[Path]] = {}
    for csv in csv_files:
        rel = csv.relative_to(base_dir)
        top = rel.parts[0] if rel.parts else "ROOT"
        grouped.setdefault(top, []).append(csv)

    for top_folder, files in grouped.items():
        out = get_excel_output_path(schema, top_folder)
        summary_rows = []

        with pd.ExcelWriter(out, engine="openpyxl") as writer:

            # --------------------------
            # 각 CSV 시트로 저장
            # --------------------------
            for csv in files:
                rel = csv.relative_to(base_dir)
                sheet = Path(rel.stem).stem.upper()[:31]

                df = pd.read_csv(csv, compression="gzip")
                row_count = len(df)

                if row_count > 1_048_576:
                    logging.warning(
                        "Excel row limit exceeded, skip sheet | %s | rows=%d",
                        sheet,
                        row_count,
                    )
                    continue

                df.to_excel(writer, sheet_name=sheet, index=False)
                summary_rows.append({"sheet_name": sheet, "rows": row_count})

                ws = writer.book[sheet]

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                header_fill = PatternFill("solid", fgColor="D9E1F2")
                header_font = Font(bold=True)

                for col_idx in range(1, len(df.columns) + 1):
                    c = ws.cell(row=1, column=col_idx)
                    c.fill = header_fill
                    c.font = header_font

                for col_idx, col_name in enumerate(df.columns, start=1):
                    max_len = len(str(col_name))
                    for v in df.iloc[:, col_idx - 1]:
                        if pd.notna(v):
                            max_len = max(max_len, len(str(v)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(
                        int(max_len * 1.2) + 2,
                        50,
                    )

                for col_idx, col_name in enumerate(df.columns, start=1):
                    if is_integer_dtype(df[col_name]) or is_float_dtype(df[col_name]):
                        for row in range(2, ws.max_row + 1):
                            ws.cell(row=row, column=col_idx).number_format = "#,##0"

            # --------------------------
            # SUMMARY 시트
            # --------------------------
            if summary_rows:
                summary_df = pd.DataFrame(summary_rows)
                summary_df.insert(0, "no", range(1, len(summary_df) + 1))

                summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)

                ws = writer.book["SUMMARY"]

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                header_fill = PatternFill("solid", fgColor="BDD7EE")
                header_font = Font(bold=True)

                for col_idx in range(1, summary_df.shape[1] + 1):
                    c = ws.cell(row=1, column=col_idx)
                    c.fill = header_fill
                    c.font = header_font

                for col_idx, col_name in enumerate(summary_df.columns, start=1):
                    max_len = len(str(col_name))
                    for v in summary_df[col_name]:
                        if pd.notna(v):
                            max_len = max(max_len, len(str(v)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(
                        int(max_len * 1.2) + 2,
                        30,
                    )

                # SUMMARY → 시트 이동 hyperlink
                existing_sheets = set(writer.book.sheetnames)

                for row_idx in range(2, ws.max_row + 1):
                    sheet_name = ws.cell(row=row_idx, column=2).value
                    if sheet_name and sheet_name in existing_sheets:
                        ws.cell(row=row_idx, column=2).hyperlink = f"#'{sheet_name}'!A1"
                        ws.cell(row=row_idx, column=2).style = "Hyperlink"

                wb = writer.book
                wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws)))

        logging.info(
            "Excel export completed | schema=%s | folder=%s | file=%s",
            schema,
            top_folder,
            out.name,
        )


# ---------------------------------------------------------
# SQL 기준 CSV 찾기
# ---------------------------------------------------------
def csv_files_from_sql(source: str, host_name: str, sql_files: list[Path]) -> list[Path]:

    csv_files: list[Path] = []

    for sql_file in sql_files:
        rel = sql_file.relative_to(SQL_DIR / source / host_name)
        subdir = rel.parent
        table = rel.stem

        pattern = f"{table}*.csv.gz"
        for csv in (CSV_DIR / source / host_name / subdir).glob(pattern):
            csv_files.append(csv)

    return csv_files
